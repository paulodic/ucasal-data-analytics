"""
22_auditoria_indicadores.py
Script global de auditoria de indicadores.
Lee los CSVs maestros de los 3 segmentos, calcula KPIs clave,
verifica consistencia interna y genera reportes comparativos.

PROPOSITO: Verificar que los datos del pipeline son coherentes entre sí y con
los reportes PDF generados. Detecta anomalias (tasa > 30%, Fuzzy excesivo, etc.)
y compara KPIs entre segmentos.

SALIDA (output_dir = outputs/General/Auditoria_Indicadores/):
  - Auditoria_Indicadores.xlsx    -> Excel con 5 hojas (KPIs, canales, checks, alertas)
  - Auditoria_Indicadores.pdf     -> Informe visual con KPIs y tabla de checks
  - Auditoria_Indicadores.md      -> Documentacion textual con resumen ejecutivo
"""
import pandas as pd
import numpy as np
import os
import sys
from datetime import datetime
from causal_utils import make_pk as _make_pk_shared
from fpdf import FPDF

# ============================================================
# CONFIGURACIÓN DE RUTAS Y CONSTANTES
# ============================================================
BASE_DIR = r"h:\Test-Antigravity\marketing_report"
DB_DIR   = os.path.join(BASE_DIR, "outputs", "Data_Base")
OUT_DIR  = os.path.join(BASE_DIR, "outputs", "General", "Auditoria_Indicadores")
os.makedirs(OUT_DIR, exist_ok=True)

SEGMENTOS = ["Grado_Pregrado", "Cursos", "Posgrados"]

# Inversión publicitaria (ARS, sin impuestos, período sep 2025 - feb 2026)
GOOGLE_SPEND = {
    'Grado_Pregrado': 47_387_402.90,
    'Cursos': 0.0,
    'Posgrados': 0.0,
}
# Facebook totales por segmento (de 20_presupuesto_roi.py)
FB_SPEND = {
    'Grado_Pregrado': 115_200_000.0,   # ~84.3% del total FB
    'Cursos':           7_100_000.0,   # ~5.2%
    'Posgrados':       14_500_000.0,   # ~10.6%
}

COHORT_START = {
    'Grado_Pregrado': '2025-09-01',   # inicio cohorte Ingreso 2026
    'Cursos': None,                    # sin restriccion de inicio
    'Posgrados': None,                 # sin restriccion de inicio
}

# ============================================================
# HELPERS
# ============================================================

def safe_div(a, b):
    return a / b if b and b != 0 else 0.0


def get_max_insc_date(df_insc: pd.DataFrame):
    """Devuelve el max(Insc_Fecha Pago) filtrado <= hoy."""
    ts = pd.Timestamp.now()
    for col in ['Insc_Fecha Pago', 'Fecha Pago']:
        if col in df_insc.columns:
            d = pd.to_datetime(df_insc[col], format='mixed', errors='coerce')
            d = d[d <= ts]
            if not d.isna().all():
                return d.max()
    return ts


def classify_match(v):
    s = str(v)
    if 'Exacto' in s: return 'Exacto'
    if 'Fuzzy'  in s: return 'Fuzzy'
    return 'Sin_Match'


def make_pk(df: pd.DataFrame) -> pd.Series:
    """Delega a la función compartida en causal_utils (DNI > Email > Tel > Cel > idx)."""
    return _make_pk_shared(df)


def classify_fuente(df: pd.DataFrame) -> pd.Series:
    """
    Clasifica cada lead en canal: Google | Facebook | Bot | Organico | Otro.
    Prioridad: UTM source/medium > FuenteLead.
    """
    fuente = pd.Series('Otro', index=df.index)

    # UTM-based
    utm_cols = [c for c in df.columns if c.lower().startswith('utm')]
    utm_text = pd.Series('', index=df.index)
    for c in utm_cols:
        utm_text = utm_text + ' ' + df[c].astype(str).str.lower().fillna('')

    google_kw = ['google', 'cpc', 'adwords', 'pmax', 'search', 'display', 'youtube', 'gdn']
    fb_kw     = ['facebook', 'fb', 'instagram', 'ig', 'meta']

    mask_g = utm_text.str.contains('|'.join(google_kw), regex=True)
    mask_f = utm_text.str.contains('|'.join(fb_kw), regex=True)

    fuente[mask_g] = 'Google'
    fuente[mask_f] = 'Facebook'

    # FuenteLead override donde UTM no clasifica
    if 'FuenteLead' in df.columns:
        fl = pd.to_numeric(df['FuenteLead'], errors='coerce')
        mask_bot = (fl == 907) & (fuente == 'Otro')
        mask_fb  = (fl == 18) & (fuente == 'Otro')
        fuente[mask_bot] = 'Bot'
        fuente[mask_fb]  = 'Facebook'

    # Orgánico: sin UTM, sin FuenteLead clave, Match existe
    mask_organico = (fuente == 'Otro') & (utm_text.str.strip() == '')
    fuente[mask_organico] = 'Organico'

    return fuente


# ============================================================
# PROCESAMIENTO POR SEGMENTO
# ============================================================
# Para cada segmento: carga CSV, calcula KPIs, verifica consistencia, detecta alertas

all_kpis       = []   # lista de dicts, uno por segmento
all_canales    = []   # lista de dicts canal×segmento
all_checks     = []   # lista de checks de consistencia
all_alertas    = []   # alertas por valores fuera de rango

for seg in SEGMENTOS:
    leads_path = os.path.join(DB_DIR, seg, "reporte_marketing_leads_completos.csv")
    insc_path  = os.path.join(DB_DIR, seg, "reporte_marketing_inscriptos_origenes.csv")

    if not os.path.exists(leads_path):
        print(f"[!] No existe CSV para {seg} — omitiendo.")
        continue

    print(f"\n{'='*55}")
    print(f"  Auditando: {seg}")
    print(f"{'='*55}")

    # Carga de datos: los CSVs maestros son grandes; se cargan completos para calcular métricas
    df_raw  = pd.read_csv(leads_path,  low_memory=False)
    df_insc = pd.read_csv(insc_path,   low_memory=False) if os.path.exists(insc_path) else pd.DataFrame()
    print(f"  Leads raw: {len(df_raw):,}  |  Inscriptos: {len(df_insc):,}")

    # Clasificación del tipo de match
    df_raw['_mc'] = df_raw['Match_Tipo'].apply(classify_match)

    # Clave de deduplicación (DNI > Correo > ID Consulta)
    df_raw['_pk'] = make_pk(df_raw)

    # Parsear fecha de creación del lead (D/M/YYYY en los CSVs)
    df_raw['_fecha'] = pd.to_datetime(
        df_raw.get('Consulta: Fecha de creación', pd.Series(dtype=str)),
        format='mixed', dayfirst=True, errors='coerce'
    )

    # Fecha máxima de inscripciones (límite superior de la ventana de análisis)
    max_insc_ts = get_max_insc_date(df_insc) if not df_insc.empty else pd.Timestamp.now()

    # Totales históricos (sin restricción de fechas)
    n_raw    = len(df_raw)
    n_exacto = (df_raw['_mc'] == 'Exacto').sum()
    n_fuzzy  = (df_raw['_mc'] == 'Fuzzy').sum()
    n_sinmatch = (df_raw['_mc'] == 'Sin_Match').sum()

    # Dedup priorizado (Exacto > Fuzzy > Sin_Match)
    prio = {'Exacto': 0, 'Fuzzy': 1, 'Sin_Match': 2}
    df_sorted = df_raw.assign(_prio=df_raw['_mc'].map(prio)).sort_values('_prio')
    df_dedup  = df_sorted.drop_duplicates(subset='_pk')
    n_dedup   = len(df_dedup)

    n_exacto_dedup = (df_dedup['_mc'] == 'Exacto').sum()
    n_fuzzy_dedup  = (df_dedup['_mc'] == 'Fuzzy').sum()
    n_sin_dedup    = (df_dedup['_mc'] == 'Sin_Match').sum()

    # Desglose por tipo de match exacto
    n_dni  = int((df_dedup['Match_Tipo'] == 'Exacto (DNI)').sum())
    n_email = int((df_dedup['Match_Tipo'] == 'Exacto (Email)').sum())
    n_tel  = int((df_dedup['Match_Tipo'] == 'Exacto (Teléfono)').sum())
    n_cel  = int((df_dedup['Match_Tipo'] == 'Exacto (Celular)').sum())

    # Ventana de conversión: período relevante para calcular tasa de conversión
    start = COHORT_START.get(seg)
    if start:
        mask_ventana = (df_raw['_fecha'] >= start) & (df_raw['_fecha'] <= max_insc_ts)
    else:
        mask_ventana = df_raw['_fecha'] <= max_insc_ts

    mask_v_sorted   = mask_ventana.reindex(df_sorted.index, fill_value=False)
    df_ventana      = df_sorted[mask_v_sorted].drop_duplicates(subset='_pk')
    n_ventana        = len(df_ventana)
    n_conv_ventana   = (df_ventana['_mc'] == 'Exacto').sum()
    tasa_conv        = safe_div(n_conv_ventana, n_ventana) * 100

    print(f"  Dedup historico: {n_dedup:,}  |  Ventana: {n_ventana:,}  |  Conv: {n_conv_ventana:,}  ({tasa_conv:.2f}%)")

    # Clasificación por canal de marketing (Google | Facebook | Bot | Organico | Otro)
    df_raw['_canal'] = classify_fuente(df_raw)
    df_dedup2 = df_sorted.drop_duplicates(subset='_pk').copy()
    df_dedup2['_canal'] = df_raw.reindex(df_dedup2.index)['_canal'].values

    canal_vol  = df_dedup2.groupby('_canal').size().rename('Personas_Dedup')
    canal_conv = df_dedup2[df_dedup2['_mc'] == 'Exacto'].groupby('_canal').size().rename('Inscriptos')
    df_canal   = pd.concat([canal_vol, canal_conv], axis=1).fillna(0).astype(int)
    df_canal['Tasa_%'] = (df_canal['Inscriptos'] / df_canal['Personas_Dedup'] * 100).round(2)
    df_canal.loc[df_canal['Personas_Dedup'] == 0, 'Tasa_%'] = 0.0
    df_canal = df_canal.reset_index().rename(columns={'_canal': 'Canal'})
    df_canal.insert(0, 'Segmento', seg)
    all_canales.append(df_canal)

    # Métricas de inversión publicitaria: CPL, CPA y ROI por canal
    g_spend = GOOGLE_SPEND.get(seg, 0.0)
    f_spend = FB_SPEND.get(seg, 0.0)

    # Leads por canal (dedup ventana)
    df_ventana2 = df_sorted[mask_v_sorted].drop_duplicates(subset='_pk').copy()
    df_ventana2['_canal'] = classify_fuente(df_raw.reindex(df_ventana2.index))

    n_leads_google = (df_ventana2['_canal'] == 'Google').sum()
    n_leads_fb     = (df_ventana2['_canal'] == 'Facebook').sum()

    n_conv_google  = ((df_ventana2['_canal'] == 'Google') & (df_ventana2['_mc'] == 'Exacto')).sum()
    n_conv_fb      = ((df_ventana2['_canal'] == 'Facebook') & (df_ventana2['_mc'] == 'Exacto')).sum()

    cpl_google = safe_div(g_spend, n_leads_google)
    cpa_google = safe_div(g_spend, n_conv_google)
    cpl_fb     = safe_div(f_spend, n_leads_fb)
    cpa_fb     = safe_div(f_spend, n_conv_fb)

    # Ingresos atribuidos (Insc_Haber)
    rev_google, rev_fb = 0.0, 0.0
    if 'Insc_Haber' in df_ventana2.columns:
        exactos_v2 = df_ventana2[df_ventana2['_mc'] == 'Exacto']
        rev_google = float(exactos_v2[exactos_v2['_canal'] == 'Google']['Insc_Haber']
                           .apply(lambda x: pd.to_numeric(str(x).replace(',', '.'), errors='coerce'))
                           .fillna(0).sum())
        rev_fb     = float(exactos_v2[exactos_v2['_canal'] == 'Facebook']['Insc_Haber']
                           .apply(lambda x: pd.to_numeric(str(x).replace(',', '.'), errors='coerce'))
                           .fillna(0).sum())

    roi_google = safe_div(rev_google - g_spend, g_spend) * 100 if g_spend > 0 else None
    roi_fb     = safe_div(rev_fb - f_spend, f_spend) * 100 if f_spend > 0 else None

    # Totales en la base de inscriptos (perspectiva del sistema contable)
    n_insc_base   = len(df_insc) if not df_insc.empty else 0
    n_insc_exacto = 0
    n_insc_directo= 0
    if not df_insc.empty and 'Match_Tipo' in df_insc.columns:
        n_insc_exacto  = df_insc['Match_Tipo'].astype(str).str.contains('Exacto').sum()
        n_insc_directo = (df_insc['Match_Tipo'] == 'No (Solo Inscripto Directo)').sum()

    # --- KPIs dict ---
    kpi = {
        'Segmento':               seg,
        'Leads_Raw':              n_raw,
        'Exacto_Raw':             int(n_exacto),
        'Fuzzy_Raw':              int(n_fuzzy),
        'Sin_Match_Raw':          int(n_sinmatch),
        'Check_Raw_Sum':          int(n_exacto + n_fuzzy + n_sinmatch),  # debe == Leads_Raw
        'Personas_Dedup':         n_dedup,
        'Exacto_Dedup':           int(n_exacto_dedup),
        'Exacto_DNI':             n_dni,
        'Exacto_Email':           n_email,
        'Exacto_Telefono':        n_tel,
        'Exacto_Celular':         n_cel,
        'Fuzzy_Dedup':            int(n_fuzzy_dedup),
        'Sin_Match_Dedup':        int(n_sin_dedup),
        'Check_Dedup_Sum':        int(n_exacto_dedup + n_fuzzy_dedup + n_sin_dedup),  # debe == Personas_Dedup
        'Ventana_Inicio':         start if start else 'Todos',
        'Ventana_Fin':            max_insc_ts.strftime('%Y-%m-%d'),
        'Leads_Ventana_Dedup':    n_ventana,
        'Conv_Ventana_Exacto':    int(n_conv_ventana),
        'Tasa_Conv_%':            round(tasa_conv, 4),
        'Inscriptos_Base':        n_insc_base,
        'Inscriptos_Exacto_Base': n_insc_exacto,
        'Inscriptos_Directo_Base':n_insc_directo,
        'Leads_Google_Ventana':   int(n_leads_google),
        'Conv_Google_Ventana':    int(n_conv_google),
        'Leads_Facebook_Ventana': int(n_leads_fb),
        'Conv_Facebook_Ventana':  int(n_conv_fb),
        'Google_Spend_ARS':       g_spend,
        'Facebook_Spend_ARS':     f_spend,
        'CPL_Google':             round(cpl_google, 2),
        'CPA_Google':             round(cpa_google, 2),
        'CPL_Facebook':           round(cpl_fb, 2),
        'CPA_Facebook':           round(cpa_fb, 2),
        'Rev_Atrib_Google':       round(rev_google, 2),
        'ROI_Google_%':           round(roi_google, 2) if roi_google is not None else None,
        'Rev_Atrib_Facebook':     round(rev_fb, 2),
        'ROI_Facebook_%':         round(roi_fb, 2) if roi_fb is not None else None,
    }
    all_kpis.append(kpi)

    # ============================================================
    # CHECKS DE CONSISTENCIA (detecta errores en el pipeline)
    # ============================================================
    checks_seg = []

    # 1. La suma de exacto + fuzzy + sin_match debe ser igual al total raw
    diff_raw = abs(kpi['Check_Raw_Sum'] - n_raw)
    checks_seg.append({
        'Segmento': seg, 'Check': 'Exacto+Fuzzy+SinMatch == Leads_Raw',
        'Esperado': n_raw, 'Obtenido': kpi['Check_Raw_Sum'],
        'Diferencia': diff_raw, 'OK': 'SI' if diff_raw == 0 else 'NO'
    })

    # 2. Suma dedup
    diff_dedup = abs(kpi['Check_Dedup_Sum'] - n_dedup)
    checks_seg.append({
        'Segmento': seg, 'Check': 'Exacto+Fuzzy+SinMatch_dedup == Personas_Dedup',
        'Esperado': n_dedup, 'Obtenido': kpi['Check_Dedup_Sum'],
        'Diferencia': diff_dedup, 'OK': 'SI' if diff_dedup == 0 else 'NO'
    })

    # 3. Ventana <= Dedup histórico
    checks_seg.append({
        'Segmento': seg, 'Check': 'Leads_Ventana_Dedup <= Personas_Dedup',
        'Esperado': f'<= {n_dedup}', 'Obtenido': n_ventana,
        'Diferencia': max(0, n_ventana - n_dedup), 'OK': 'SI' if n_ventana <= n_dedup else 'NO'
    })

    # 4. Conv ventana <= Leads ventana
    checks_seg.append({
        'Segmento': seg, 'Check': 'Conv_Ventana <= Leads_Ventana_Dedup',
        'Esperado': f'<= {n_ventana}', 'Obtenido': n_conv_ventana,
        'Diferencia': max(0, n_conv_ventana - n_ventana),
        'OK': 'SI' if n_conv_ventana <= n_ventana else 'NO'
    })

    # 5. Tasa conversión razonable (0–30%)
    checks_seg.append({
        'Segmento': seg, 'Check': 'Tasa_Conv en rango [0%, 30%]',
        'Esperado': '0%-30%', 'Obtenido': f'{tasa_conv:.2f}%',
        'Diferencia': '', 'OK': 'SI' if 0 <= tasa_conv <= 30 else 'ALERTA'
    })

    # 6. Inscriptos base >= Conv_Ventana (inscriptos base contiene otros canales también)
    checks_seg.append({
        'Segmento': seg, 'Check': 'Inscriptos_Base >= Conv_Ventana_Exacto',
        'Esperado': f'>= {n_conv_ventana}', 'Obtenido': n_insc_base,
        'Diferencia': max(0, n_conv_ventana - n_insc_base),
        'OK': 'SI' if n_insc_base >= n_conv_ventana else 'ALERTA'
    })

    all_checks.extend(checks_seg)

    # ============================================================
    # ALERTAS (valores fuera de rangos esperados)
    # ============================================================
    if tasa_conv > 30:
        all_alertas.append({'Segmento': seg, 'Indicador': 'Tasa_Conv_%',
                             'Valor': f'{tasa_conv:.2f}%', 'Alerta': 'Tasa > 30% — revisar ventana cohorte'})
    if tasa_conv < 0.5 and n_ventana > 100:
        all_alertas.append({'Segmento': seg, 'Indicador': 'Tasa_Conv_%',
                             'Valor': f'{tasa_conv:.2f}%', 'Alerta': 'Tasa < 0.5% con muestra grande — verificar matching'})
    if n_fuzzy > n_exacto * 3:
        all_alertas.append({'Segmento': seg, 'Indicador': 'Fuzzy_Raw',
                             'Valor': int(n_fuzzy), 'Alerta': 'Fuzzy > 3x Exacto — muchos matcheos dudosos'})
    if n_insc_base > 0 and n_insc_exacto == 0:
        all_alertas.append({'Segmento': seg, 'Indicador': 'Inscriptos_Exacto_Base',
                             'Valor': 0, 'Alerta': 'Sin inscriptos exactos en base — columna Match_Tipo puede estar mal'})
    if g_spend > 0 and n_leads_google == 0:
        all_alertas.append({'Segmento': seg, 'Indicador': 'Leads_Google_Ventana',
                             'Valor': 0, 'Alerta': 'Inversión Google > 0 pero 0 leads atribuidos — revisar UTMs'})

    print(f"  Checks: {len(checks_seg)} | Alertas: {len([a for a in all_alertas if a['Segmento'] == seg])}")


# ============================================================
# CONSTRUIR DATAFRAMES DE SALIDA
# ============================================================

df_kpis    = pd.DataFrame(all_kpis)
df_checks  = pd.DataFrame(all_checks)
df_alertas = pd.DataFrame(all_alertas) if all_alertas else pd.DataFrame(
    columns=['Segmento', 'Indicador', 'Valor', 'Alerta'])
df_canales = pd.concat(all_canales, ignore_index=True) if all_canales else pd.DataFrame()

# Trasponer KPIs para visualización vertical (Métrica × Segmento)
metric_cols = [c for c in df_kpis.columns if c != 'Segmento']
df_kpis_v   = df_kpis.set_index('Segmento')[metric_cols].T.reset_index()
df_kpis_v.columns = ['Metrica'] + list(df_kpis_v.columns[1:])

# Resumen global (suma / promedio según tipo)
df_kpis_v['TOTAL / PROM'] = ''
numeric_rows = df_kpis_v['Metrica'].str.contains(
    'Leads|Conv|Inscriptos|Personas|Exacto|Fuzzy|Sin_Match|Spend|Rev_Atrib|Check',
    regex=True, case=False, na=False)
for idx, row in df_kpis_v.iterrows():
    try:
        vals = pd.to_numeric(df_kpis_v.loc[idx, SEGMENTOS], errors='coerce')
        if numeric_rows[idx]:
            df_kpis_v.at[idx, 'TOTAL / PROM'] = vals.sum()
        elif 'Tasa' in row['Metrica'] or 'CPL' in row['Metrica'] or 'CPA' in row['Metrica']:
            df_kpis_v.at[idx, 'TOTAL / PROM'] = f"prom {vals.mean():.2f}"
    except Exception:
        pass

# ============================================================
# EXPORTAR EXCEL
# ============================================================
out_file = os.path.join(OUT_DIR, "Auditoria_Indicadores.xlsx")
print(f"\nEscribiendo Excel: {out_file}")

with pd.ExcelWriter(out_file, engine='openpyxl') as writer:

    # --- Hoja 1: KPIs por segmento (vertical) ---
    df_kpis_v.to_excel(writer, sheet_name='KPIs_Por_Segmento', index=False)

    # --- Hoja 2: Canales ---
    if not df_canales.empty:
        df_canales.to_excel(writer, sheet_name='Canal_X_Segmento', index=False)

    # --- Hoja 3: Checks de consistencia ---
    df_checks.to_excel(writer, sheet_name='Checks_Consistencia', index=False)

    # --- Hoja 4: Alertas ---
    df_alertas.to_excel(writer, sheet_name='Alertas', index=False)

    # --- Hoja 5: KPIs horizontal (raw) ---
    df_kpis.to_excel(writer, sheet_name='KPIs_Horizontal', index=False)

# --- Aplicar formato mínimo con openpyxl ---
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
WHITE_FONT  = Font(bold=True, color="FFFFFF", size=9)
BOLD_FONT   = Font(bold=True, size=9)
STD_FONT    = Font(size=9)

wb = load_workbook(out_file)

def format_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30

def autofit_columns(ws, min_w=12, max_w=40):
    for col_cells in ws.columns:
        length = max(len(str(c.value or '')) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_w, min(length + 3, max_w))

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    format_header_row(ws)
    ws.freeze_panes = 'B2' if sheet_name == 'KPIs_Por_Segmento' else 'A2'
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)

    # Colorear checks OK/NO
    if sheet_name == 'Checks_Consistencia':
        ok_col = None
        for idx, cell in enumerate(ws[1], start=1):
            if str(cell.value) == 'OK':
                ok_col = idx
                break
        if ok_col:
            for row in ws.iter_rows(min_row=2):
                val = str(row[ok_col - 1].value or '')
                if val == 'SI':
                    row[ok_col - 1].fill = GREEN_FILL
                elif val == 'NO':
                    row[ok_col - 1].fill = RED_FILL
                elif val == 'ALERTA':
                    row[ok_col - 1].fill = YELLOW_FILL

    # Colorear alertas
    if sheet_name == 'Alertas':
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = YELLOW_FILL

wb.save(out_file)
size_mb = os.path.getsize(out_file) / 1_048_576
print(f"\n-> Excel generado: {out_file}")
print(f"   Tamano: {size_mb:.1f} MB")
print(f"   Hojas: KPIs_Por_Segmento | Canal_X_Segmento | Checks_Consistencia | Alertas | KPIs_Horizontal")
n_ok    = (df_checks['OK'] == 'SI').sum()
n_fail  = (df_checks['OK'] != 'SI').sum()
print(f"\nRESUMEN CHECKS: {n_ok} OK | {n_fail} con problemas")
if not df_alertas.empty:
    print(f"ALERTAS GENERADAS: {len(df_alertas)}")
    for _, a in df_alertas.iterrows():
        print(f"  [{a['Segmento']}] {a['Indicador']}: {a['Alerta']}")

# ============================================================
# GENERAR PDF: RESUMEN VISUAL DE LA AUDITORÍA
# ============================================================
# Helvetica solo soporta latin-1: reemplazar caracteres fuera de rango
def safe_pdf_text(s):
    """Convierte texto a latin-1 safe para fpdf2 Helvetica."""
    return (str(s)
            .replace('\u2014', '-')   # em dash
            .replace('\u2013', '-')   # en dash
            .replace('\u2019', "'")   # comilla curva
            .replace('\u201c', '"')
            .replace('\u201d', '"')
            .encode('latin-1', errors='replace').decode('latin-1'))

print("\nGenerando PDF de auditoria...")
pdf = FPDF('L')  # Landscape para que las tablas no se corten
pdf.add_page()

# Encabezado
pdf.set_font('Helvetica', 'B', 15)
pdf.cell(0, 10, 'Auditoria de Indicadores - UCASAL Marketing', ln=True, align='C')
pdf.set_font('Helvetica', 'I', 9)
pdf.cell(0, 6, f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  '
               f'Segmentos: {", ".join(SEGMENTOS)}', ln=True, align='C')
pdf.ln(5)

# KPIs por segmento
pdf.set_font('Helvetica', 'B', 11)
pdf.cell(0, 8, '1. KPIs Clave por Segmento', ln=True)
pdf.set_font('Helvetica', '', 9)
# Métricas a mostrar en el PDF (subconjunto de todas las métricas del Excel)
metricas_pdf = [
    'Leads_Raw', 'Personas_Dedup', 'Exacto_Dedup',
    'Exacto_DNI', 'Exacto_Email', 'Exacto_Telefono', 'Exacto_Celular',
    'Leads_Ventana_Dedup', 'Conv_Ventana_Exacto', 'Tasa_Conv_%',
    'Google_Spend_ARS', 'Facebook_Spend_ARS', 'CPL_Google', 'CPA_Google',
    'CPL_Facebook', 'CPA_Facebook', 'ROI_Google_%', 'ROI_Facebook_%',
]
# Encabezado de tabla
col_w_m = 80
col_w_v = 60
pdf.set_font('Helvetica', 'B', 8)
pdf.set_fill_color(31, 78, 121)
pdf.set_text_color(255, 255, 255)
pdf.cell(col_w_m, 7, 'Metrica', border=1, fill=True)
for seg in SEGMENTOS:
    pdf.cell(col_w_v, 7, seg.replace('_', ' '), border=1, fill=True, align='C')
pdf.ln()
pdf.set_text_color(0, 0, 0)

# Filas de métricas
for i, metrica in enumerate(metricas_pdf):
    fill = (i % 2 == 0)
    if fill:
        pdf.set_fill_color(235, 245, 255)
    else:
        pdf.set_fill_color(255, 255, 255)
    pdf.set_font('Helvetica', '', 8)
    pdf.cell(col_w_m, 6, metrica, border=1, fill=fill)
    for kpi in all_kpis:
        if kpi['Segmento'] in SEGMENTOS:
            val = kpi.get(metrica, '')
            if isinstance(val, float):
                val_str = f"{val:,.2f}" if val < 1_000_000 else f"{val:,.0f}"
            elif isinstance(val, int):
                val_str = f"{val:,}"
            else:
                val_str = str(val) if val is not None else ''
            pdf.cell(col_w_v, 6, val_str, border=1, fill=fill, align='R')
    pdf.ln()

pdf.ln(5)

# Checks de consistencia
pdf.add_page()
pdf.set_font('Helvetica', 'B', 11)
pdf.cell(0, 8, '2. Checks de Consistencia del Pipeline', ln=True)
pdf.set_font('Helvetica', 'B', 8)
pdf.set_fill_color(31, 78, 121)
pdf.set_text_color(255, 255, 255)
chk_ws = [40, 110, 35, 35, 25]
chk_hs = ['Segmento', 'Check', 'Esperado', 'Obtenido', 'OK']
for h, w in zip(chk_hs, chk_ws):
    pdf.cell(w, 7, h, border=1, fill=True, align='C')
pdf.ln()
pdf.set_text_color(0, 0, 0)
for i, (_, row) in enumerate(df_checks.iterrows()):
    pdf.set_font('Helvetica', '', 7)
    ok_val = str(row.get('OK', ''))
    if ok_val == 'SI':
        pdf.set_fill_color(198, 239, 206)   # verde
    elif ok_val == 'NO':
        pdf.set_fill_color(255, 199, 206)   # rojo
    else:
        pdf.set_fill_color(255, 235, 156)   # amarillo (ALERTA)
    vals = [
        safe_pdf_text(row.get('Segmento', '')),
        safe_pdf_text(str(row.get('Check', ''))[:75]),
        safe_pdf_text(str(row.get('Esperado', ''))[:25]),
        safe_pdf_text(str(row.get('Obtenido', ''))[:25]),
        ok_val,
    ]
    for v, w in zip(vals, chk_ws):
        pdf.cell(w, 6, v, border=1, fill=True)
    pdf.ln()
    pdf.set_fill_color(255, 255, 255)

# Alertas (si existen)
if not df_alertas.empty:
    pdf.ln(4)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(0, 8, '3. Alertas Detectadas', ln=True)
    pdf.set_font('Helvetica', '', 9)
    for _, a in df_alertas.iterrows():
        pdf.set_fill_color(255, 235, 156)
        pdf.cell(0, 6, safe_pdf_text(f"[{a['Segmento']}] {a['Indicador']}: {a['Alerta']}"), ln=True, fill=True)
    pdf.set_fill_color(255, 255, 255)

pdf_path = os.path.join(OUT_DIR, 'Auditoria_Indicadores.pdf')
pdf.output(pdf_path)
print(f"-> PDF generado: {pdf_path}")

# ============================================================
# DOCUMENTACIÓN MARKDOWN
# ============================================================
md_lines = [
    "# Auditoria de Indicadores - UCASAL Marketing\n",
    f"**Generado:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  \n",
    f"**Script:** `22_auditoria_indicadores.py`  \n",
    f"**Checks:** {n_ok} OK | {n_fail} con problemas  \n",
    f"**Alertas:** {len(df_alertas)}  \n\n",
    "## KPIs por Segmento\n\n",
    df_kpis_v[['Metrica'] + SEGMENTOS].to_markdown(index=False),
    "\n\n## Checks de Consistencia\n\n",
    df_checks.to_markdown(index=False),
]
if not df_alertas.empty:
    md_lines += ["\n\n## Alertas\n\n", df_alertas.to_markdown(index=False)]
md_lines += [
    "\n\n## Archivos de Salida\n",
    "| Archivo | Descripcion |\n|---|---|\n",
    f"| `Auditoria_Indicadores.xlsx` | Excel con 5 hojas (KPIs, canales, checks, alertas) |\n",
    f"| `Auditoria_Indicadores.pdf` | Informe visual con KPIs y checks |\n",
    f"| `Auditoria_Indicadores.md` | Este archivo |\n",
]
md_path = os.path.join(OUT_DIR, 'Auditoria_Indicadores.md')
with open(md_path, 'w', encoding='utf-8') as f:
    f.writelines(md_lines)
print(f"-> MD generado: {md_path}")
print("\nProceso finalizado.")
